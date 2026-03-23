"""Load curated agent rows from resources/agent-list.xlsx and merge optional links."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parent.parent.parent
_XLSX_PATH = _ROOT / "resources" / "agent-list.xlsx"
_LINKS_PATH = _ROOT / "resources" / "agent-links.json"
_LOCALE_EN_PATH = _ROOT / "resources" / "agent-locale.en.json"

# Excel column header -> API field (stable English keys)
_HEADER_TO_FIELD: Dict[str, str] = {
    "名称": "name",
    "公司/组织": "organization",
    "类型": "agent_type",
    "分类层级": "category_tier",
    "是否开源": "open_source",
    "主要场景": "scenarios",
    "特色/定位": "positioning",
    "专注领域": "focus_areas",
    "技术优势": "technical_strengths",
    "局限/注意点": "limitations",
    "推荐级别": "recommendation_level",
    "备注": "notes",
}


def _load_links() -> Dict[str, Dict[str, str]]:
    if not _LINKS_PATH.is_file():
        return {}
    with _LINKS_PATH.open(encoding="utf-8") as f:
        raw = json.load(f)
    return raw if isinstance(raw, dict) else {}


def _load_locale_en() -> Dict[str, Dict[str, Any]]:
    if not _LOCALE_EN_PATH.is_file():
        return {}
    with _LOCALE_EN_PATH.open(encoding="utf-8") as f:
        raw = json.load(f)
    return raw if isinstance(raw, dict) else {}


def _apply_en_overrides(rec: Dict[str, Any], overrides: Dict[str, Any]) -> None:
    for key, val in overrides.items():
        if key not in rec or val is None:
            continue
        s = str(val).strip()
        if s:
            rec[key] = s


def load_agents(locale: str = "zh") -> List[Dict[str, Any]]:
    """Return agent records with integer id (1-based row order), links, and optional EN copy."""
    if not _XLSX_PATH.is_file():
        return []
    links = _load_links()
    wb = load_workbook(_XLSX_PATH, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header_row = rows[0]
    col_index: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip()
        if key in _HEADER_TO_FIELD:
            col_index[key] = i
    out: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows[1:], start=1):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        rec: Dict[str, Any] = {"id": idx}
        for zh, field in _HEADER_TO_FIELD.items():
            j = col_index.get(zh)
            val = row[j] if j is not None and j < len(row) else None
            rec[field] = "" if val is None else str(val).strip()
        name = rec.get("name") or ""
        extra = links.get(name) if isinstance(links.get(name), dict) else {}
        rec["website"] = (extra or {}).get("website") or None
        rec["documentation"] = (extra or {}).get("documentation") or None
        out.append(rec)
    lang = (locale or "zh").strip().lower()
    if lang.startswith("en"):
        en_map = _load_locale_en()
        for rec in out:
            ov = en_map.get(str(rec["id"]))
            if isinstance(ov, dict):
                _apply_en_overrides(rec, ov)
    return out
